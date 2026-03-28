#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
import psycopg
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SERVICE_LOOP_SQL = """
with latest_processes as (
  select distinct on (p.id)
    p.id,
    p.state_code,
    p.version,
    p.created_at,
    p.json::jsonb as j
  from processes p
  order by p.id, p.created_at desc
),
process_meta as (
  select
    lp.id as process_id,
    lp.state_code,
    lp.version as process_version,
    coalesce(
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,1,#text}',
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,0,#text}'
    ) as process_name,
    coalesce(
      lp.j #>> '{processDataSet,processInformation,geography,locationOfOperationSupplyOrProduction,@location}',
      lp.j #>> '{processDataSet,processInformation,geography,subLocationOfOperationSupplyOrProduction,@subLocation}',
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,locationOfOperationSupplyOrProduction}'
    ) as location,
    lp.j #>> '{processDataSet,quantitativeReference,referenceToReferenceFlow}' as reference_exchange_internal_id,
    case
      when jsonb_typeof(lp.j #> '{processDataSet,exchanges,exchange}') = 'array' then lp.j #> '{processDataSet,exchanges,exchange}'
      when lp.j #> '{processDataSet,exchanges,exchange}' is null then '[]'::jsonb
      else jsonb_build_array(lp.j #> '{processDataSet,exchanges,exchange}')
    end as exchanges
  from latest_processes lp
),
exchanges as (
  select
    pm.process_id,
    pm.process_name,
    pm.location,
    pm.state_code,
    pm.process_version,
    pm.reference_exchange_internal_id,
    ex.value #>> '{@dataSetInternalID}' as exchange_internal_id,
    ex.value #>> '{exchangeDirection}' as direction,
    ex.value #>> '{referenceToFlowDataSet,@refObjectId}' as flow_id,
    coalesce(
      ex.value #>> '{referenceToFlowDataSet,common:shortDescription,#text}',
      ex.value #>> '{referenceToFlowDataSet,shortDescription,#text}'
    ) as flow_name,
    trim(replace(replace(coalesce(ex.value #>> '{resultingAmount}', ex.value #>> '{meanAmount}', ''), chr(160), ''), ',', '')) as amount_text,
    coalesce(
      ex.value #>> '{generalComment,1,#text}',
      ex.value #>> '{generalComment,0,#text}'
    ) as comment_text
  from process_meta pm
  cross join lateral jsonb_array_elements(pm.exchanges) ex(value)
)
select
  i.process_id,
  i.process_name,
  i.location,
  i.state_code,
  i.process_version,
  i.flow_id,
  i.flow_name,
  i.reference_exchange_internal_id,
  i.exchange_internal_id as input_exchange_internal_id,
  o.exchange_internal_id as output_exchange_internal_id,
  (o.exchange_internal_id = i.reference_exchange_internal_id) as output_is_reference,
  i.input_amount_text,
  o.output_amount_text,
  i.comment_text as input_comment,
  o.comment_text as output_comment,
  'exact_same_flow_input_output_text' as suspicion_tag
from (
  select *, amount_text as input_amount_text from exchanges where direction = 'Input'
) i
join (
  select *, amount_text as output_amount_text from exchanges where direction = 'Output'
) o
  on i.process_id = o.process_id
 and i.flow_id = o.flow_id
where i.input_amount_text <> ''
  and i.input_amount_text = o.output_amount_text
order by
  (o.exchange_internal_id = i.reference_exchange_internal_id) desc,
  i.input_amount_text desc,
  i.process_name nulls last
"""


PN_PM_SQL = """
with latest_processes as (
  select distinct on (p.id)
    p.id,
    p.state_code,
    p.version,
    p.created_at,
    p.json::jsonb as j
  from processes p
  order by p.id, p.created_at desc
),
latest_flows as (
  select distinct on (f.id)
    f.id,
    f.version,
    f.created_at,
    f.json::jsonb as j
  from flows f
  order by f.id, f.created_at desc
),
process_meta as (
  select
    lp.id as process_id,
    lp.state_code,
    lp.version as process_version,
    coalesce(
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,1,#text}',
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,0,#text}'
    ) as process_name,
    coalesce(
      lp.j #>> '{processDataSet,processInformation,geography,locationOfOperationSupplyOrProduction,@location}',
      lp.j #>> '{processDataSet,processInformation,geography,subLocationOfOperationSupplyOrProduction,@subLocation}',
      lp.j #>> '{processDataSet,processInformation,dataSetInformation,locationOfOperationSupplyOrProduction}'
    ) as location,
    case
      when jsonb_typeof(lp.j #> '{processDataSet,exchanges,exchange}') = 'array' then lp.j #> '{processDataSet,exchanges,exchange}'
      when lp.j #> '{processDataSet,exchanges,exchange}' is null then '[]'::jsonb
      else jsonb_build_array(lp.j #> '{processDataSet,exchanges,exchange}')
    end as exchanges
  from latest_processes lp
),
process_exchanges as (
  select
    pm.process_id,
    pm.process_name,
    pm.location,
    pm.state_code,
    pm.process_version,
    ex.value #>> '{@dataSetInternalID}' as exchange_internal_id,
    ex.value #>> '{exchangeDirection}' as direction,
    ex.value #>> '{referenceToFlowDataSet,@refObjectId}' as flow_id,
    trim(replace(replace(coalesce(ex.value #>> '{resultingAmount}', ex.value #>> '{meanAmount}', ''), chr(160), ''), ',', '')) as amount_text,
    coalesce(
      ex.value #>> '{generalComment,1,#text}',
      ex.value #>> '{generalComment,0,#text}'
    ) as comment_text
  from process_meta pm
  cross join lateral jsonb_array_elements(pm.exchanges) ex(value)
),
flow_meta as (
  select
    lf.id::text as flow_id,
    coalesce(
      lf.j #>> '{flowDataSet,flowInformation,dataSetInformation,name,baseName,1,#text}',
      lf.j #>> '{flowDataSet,flowInformation,dataSetInformation,name,baseName,0,#text}'
    ) as flow_name,
    coalesce(
      lf.j #>> '{flowDataSet,flowProperties,flowProperty,referenceToFlowPropertyDataSet,common:shortDescription,#text}',
      lf.j #>> '{flowDataSet,flowProperties,flowProperty,referenceToFlowPropertyDataSet,shortDescription,#text}'
    ) as reference_property_name
  from latest_flows lf
)
select
  pe.process_id,
  pe.process_name,
  pe.location,
  pe.state_code,
  pe.process_version,
  pe.exchange_internal_id,
  pe.amount_text,
  pe.flow_id,
  fm.flow_name,
  fm.reference_property_name,
  pe.comment_text,
  case
    when pe.comment_text ilike '%PN%' and coalesce(fm.reference_property_name, '') ilike '%mass%' then 'pn_comment_but_mass_flow'
    when coalesce(fm.flow_name, '') ilike '%PM0.2%' then 'pm02_flow_hit'
    when coalesce(fm.flow_name, '') ilike '%particle%' then 'particle_flow_hit'
    else 'other_particle_suspicion'
  end as suspicion_tag
from process_exchanges pe
left join flow_meta fm on fm.flow_id = pe.flow_id
where pe.direction = 'Output'
  and (
    (pe.comment_text ilike '%PN%' and coalesce(fm.reference_property_name, '') ilike '%mass%')
    or coalesce(fm.flow_name, '') ilike '%PM0.2%'
    or coalesce(fm.flow_name, '') ilike '%particle%'
  )
order by
  case when pe.comment_text ilike '%PN%' and coalesce(fm.reference_property_name, '') ilike '%mass%' then 0 else 1 end,
  length(pe.amount_text) desc,
  pe.process_name nulls last
"""


TAG_GUIDE = [
    (
        "exact_same_flow_input_output_text",
        "同一个 process 内，同一个 flow 同时既是 Input 又是 Output，且两边 amount 文本完全一样。常见于服务流自循环/互循环的强信号。",
    ),
    (
        "pn_comment_but_mass_flow",
        "exchange 的注释写的是 PN（粒子数），但对应 flow 的 reference property 是 Mass（质量）。这是高优先级语义错配。",
    ),
    (
        "pm02_flow_hit",
        "命中了 PM0.2 相关 flow。这里是扩展排查集，不代表一定有错，但值得继续看。",
    ),
    (
        "particle_flow_hit",
        "命中了 particle 相关 flow。这里也是扩展排查集，用于扩大人工筛查范围。",
    ),
]


FIELD_GUIDE = [
    ("process_id", "问题 process 的 UUID，是后续追查和修数最重要的定位字段。"),
    ("process_name", "process 名称。"),
    ("location", "地理位置字段，便于区分同名 process。"),
    ("state_code", "process 当前版本在源表中的状态码，可用于区分已发布、草稿或其他状态数据。"),
    ("process_version", "当前命中的 process 版本。"),
    ("flow_id", "命中的 flow UUID。"),
    ("flow_name", "命中的 flow 名称。"),
    ("reference_exchange_internal_id", "如果 process 定义了 referenceToReferenceFlow，这里会显示 reference exchange 的 internal id。为空不代表没问题，只表示当前 JSON 路径上没有取到。"),
    ("input_exchange_internal_id / output_exchange_internal_id", "service-loop sheet 中，命中的输入/输出 exchange internal id。"),
    ("output_is_reference", "service-loop sheet 中，Output exchange 是否正好是 reference exchange。为 true 时风险通常更高。"),
    ("input_amount_text / output_amount_text", "service-loop sheet 中，输入/输出 amount 的原始文本。这里保留文本是为了避免被脏数据格式影响。"),
    ("amount_text", "pn_pm sheet 中，exchange 的原始 amount 文本。"),
    ("comment_text", "exchange 注释文本，例如 PN、PM、Light vehicles running 1km。"),
    ("reference_property_name", "flow 的 reference property 名称，例如 Mass。"),
    ("suspicion_tag", "这条记录为什么被命中；先看这个字段，再结合 flow_name / comment_text 解读。"),
]


SUMMARY_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
TAG_FILLS = {
    "pn_comment_but_mass_flow": PatternFill("solid", fgColor="FCE4D6"),
    "pm02_flow_hit": PatternFill("solid", fgColor="FFF2CC"),
    "particle_flow_hit": PatternFill("solid", fgColor="F4F4F4"),
    "exact_same_flow_input_output_text": PatternFill("solid", fgColor="E2F0D9"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export provider-link diagnostic process lists to a styled Excel workbook."
    )
    parser.add_argument(
        "--report-dir",
        default="reports/provider-link-diagnostics",
        help="Directory to write the Excel workbook into.",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Exact output xlsx path. If omitted, use report-dir + timestamped filename.",
    )
    parser.add_argument(
        "--filename-prefix",
        default="provider_link_problem_processes",
        help="Filename prefix when --output is not provided.",
    )
    parser.add_argument(
        "--conn-env",
        default="CONN",
        help="Primary environment variable name for database connection. Falls back to DATABASE_URL.",
    )
    return parser.parse_args()


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return
    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        os.environ.setdefault(key, value)


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def database_url(conn_env_name: str) -> str:
    root = repo_root()
    load_dotenv(root / ".env")
    db_url = os.environ.get(conn_env_name) or os.environ.get("DATABASE_URL", "")
    if not db_url:
        raise SystemExit(
            f"missing database connection; set {conn_env_name} or DATABASE_URL, or provide them in {root / '.env'}"
        )
    return db_url


def build_summary_df(service_df: pd.DataFrame, pn_pm_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = [
        {"项目": "生成时间", "说明": datetime.now().isoformat(timespec="seconds")},
        {"项目": "service_loop_row_count", "说明": int(len(service_df))},
        {
            "项目": "service_loop_distinct_process_count",
            "说明": int(service_df["process_id"].nunique()) if not service_df.empty else 0,
        },
        {"项目": "pn_pm_row_count", "说明": int(len(pn_pm_df))},
        {
            "项目": "pn_pm_distinct_process_count",
            "说明": int(pn_pm_df["process_id"].nunique()) if not pn_pm_df.empty else 0,
        },
    ]

    if not service_df.empty:
        for state_code, count in Counter(service_df["state_code"]).most_common():
            rows.append({"项目": f"service_loop_state_code::{state_code}", "说明": int(count)})
    if not pn_pm_df.empty:
        for tag, count in pn_pm_df["suspicion_tag"].value_counts().items():
            rows.append({"项目": f"pn_pm_tag::{tag}", "说明": int(count)})
        for state_code, count in Counter(pn_pm_df["state_code"]).most_common():
            rows.append({"项目": f"pn_pm_state_code::{state_code}", "说明": int(count)})

    return pd.DataFrame(rows)


def autosize_columns(ws, max_width: int = 42) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 10), max_width)


def add_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_summary_sheet(
    path: Path,
    summary_df: pd.DataFrame,
    tag_guide_df: pd.DataFrame,
    field_guide_df: pd.DataFrame,
) -> None:
    wb = load_workbook(path)
    ws = wb["summary"]
    ws.freeze_panes = "A3"

    ws["A1"] = "Provider Link 问题 Process 诊断汇总"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "本表汇总了 service-loop 和 PN/PM 语义错配两类可疑 process，并解释 tag 与关键字段含义。"
    ws["A2"].alignment = Alignment(wrap_text=True)

    stats_header_row = 3
    for cell in ws[stats_header_row]:
        cell.fill = SUMMARY_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

    tag_title_row = len(summary_df) + 6
    tag_header_row = tag_title_row + 1
    field_title_row = len(summary_df) + len(tag_guide_df) + 10
    field_header_row = field_title_row + 1
    ws[f"A{tag_title_row}"] = "Tag 含义"
    ws[f"A{field_title_row}"] = "关键字段含义"
    ws[f"A{tag_title_row}"].fill = SECTION_FILL
    ws[f"A{field_title_row}"].fill = SECTION_FILL
    ws[f"A{tag_title_row}"].font = BOLD_FONT
    ws[f"A{field_title_row}"].font = BOLD_FONT
    for cell in ws[tag_header_row]:
        cell.fill = SUMMARY_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    for cell in ws[field_header_row]:
        cell.fill = SUMMARY_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row >= stats_header_row:
                cell.border = THIN_BORDER

    autosize_columns(ws, max_width=64)
    wb.save(path)


def style_data_sheet(path: Path, sheet_name: str, tag_column: str | None = None, bool_column: str | None = None) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    tag_idx = header_map.get(tag_column) if tag_column else None
    bool_idx = header_map.get(bool_column) if bool_column else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        if tag_idx:
            tag_value = row[tag_idx - 1].value
            fill = TAG_FILLS.get(tag_value)
            if fill:
                for cell in row:
                    cell.fill = fill
        if bool_idx and row[bool_idx - 1].value is True:
            for cell in row:
                cell.font = BOLD_FONT

    autosize_columns(ws)
    add_table(ws, 1, ws.max_row, ws.max_column, f"{sheet_name.replace('-', '_')}_tbl")
    wb.save(path)


def export_workbook(output_path: Path, service_df: pd.DataFrame, pn_pm_df: pd.DataFrame) -> None:
    summary_df = build_summary_df(service_df, pn_pm_df)
    tag_guide_df = pd.DataFrame(TAG_GUIDE, columns=["tag", "中文解释"])
    field_guide_df = pd.DataFrame(FIELD_GUIDE, columns=["字段", "中文解释"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="summary", startrow=2)
        tag_guide_df.to_excel(writer, index=False, sheet_name="summary", startrow=len(summary_df) + 6)
        field_guide_df.to_excel(
            writer,
            index=False,
            sheet_name="summary",
            startrow=len(summary_df) + len(tag_guide_df) + 10,
        )
        service_df.to_excel(writer, index=False, sheet_name="service_loop_candidates")
        pn_pm_df.to_excel(writer, index=False, sheet_name="pn_pm_candidates")

    style_summary_sheet(output_path, summary_df, tag_guide_df, field_guide_df)
    style_data_sheet(
        output_path,
        "service_loop_candidates",
        tag_column="suspicion_tag",
        bool_column="output_is_reference",
    )
    style_data_sheet(output_path, "pn_pm_candidates", tag_column="suspicion_tag")


def query_dataframe(conn: psycopg.Connection, sql: str) -> pd.DataFrame:
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()
        columns = [desc.name for desc in cur.description]
    return pd.DataFrame(rows, columns=columns)


def output_path_from_args(args: argparse.Namespace) -> Path:
    root = repo_root()
    if args.output:
        return Path(args.output).expanduser().resolve()
    report_dir = (root / args.report_dir).resolve()
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return report_dir / f"{args.filename_prefix}_{ts}.xlsx"


def main() -> None:
    args = parse_args()
    db_url = database_url(args.conn_env)
    out_path = output_path_from_args(args)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    conn = psycopg.connect(db_url)
    try:
        service_df = query_dataframe(conn, SERVICE_LOOP_SQL)
        pn_pm_df = query_dataframe(conn, PN_PM_SQL)
    finally:
        conn.close()

    export_workbook(out_path, service_df, pn_pm_df)

    print(f"OUTPUT_XLSX {out_path}")
    print(f"SERVICE_LOOP_ROWS {len(service_df)}")
    print(
        "SERVICE_LOOP_PROCESSES",
        service_df["process_id"].nunique() if not service_df.empty else 0,
    )
    print(f"PN_PM_ROWS {len(pn_pm_df)}")
    print("PN_PM_PROCESSES", pn_pm_df["process_id"].nunique() if not pn_pm_df.empty else 0)


if __name__ == "__main__":
    main()
