#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
import psycopg
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ACTIVE_SNAPSHOT_SQL = """
select
  act.scope as snapshot_scope,
  act.snapshot_id,
  act.activated_at as snapshot_activated_at,
  ns.provider_matching_rule as snapshot_provider_rule,
  coalesce(ns.process_filter ->> 'selection_mode', '') as selection_mode,
  coalesce(ns.process_filter ->> 'include_user_id', '') as include_user_id,
  coalesce(
    array_to_string(
      array(
        select value::text
        from jsonb_array_elements_text(coalesce(ns.process_filter -> 'process_states', '[]'::jsonb))
      ),
      ','
    ),
    ''
  ) as process_states_label,
  coalesce((sa.coverage #>> '{matching,input_edges_total}')::bigint, 0) as input_edges_total,
  coalesce((sa.coverage #>> '{matching,matched_unique_provider}')::bigint, 0) as matched_unique_provider,
  coalesce((sa.coverage #>> '{matching,matched_multi_provider}')::bigint, 0) as matched_multi_provider,
  coalesce((sa.coverage #>> '{matching,matched_multi_resolved}')::bigint, 0) as matched_multi_resolved,
  coalesce((sa.coverage #>> '{matching,matched_multi_unresolved}')::bigint, 0) as matched_multi_unresolved,
  coalesce((sa.coverage #>> '{matching,matched_multi_fallback_equal}')::bigint, 0) as matched_multi_fallback_equal,
  coalesce((sa.coverage #>> '{matching,unmatched_no_provider}')::bigint, 0) as unmatched_no_provider,
  coalesce((sa.coverage #>> '{matching,a_input_edges_written}')::bigint, 0) as a_input_edges_written
from lca_active_snapshots act
join lca_network_snapshots ns on ns.id = act.snapshot_id
left join lca_snapshot_artifacts sa on sa.snapshot_id = act.snapshot_id
order by act.activated_at desc
limit 1
"""


DETAIL_SQL = r"""
with active_snapshot as (
  select
    act.scope as snapshot_scope,
    act.snapshot_id,
    act.activated_at as snapshot_activated_at,
    ns.provider_matching_rule as snapshot_provider_rule,
    coalesce(ns.process_filter ->> 'selection_mode', '') as selection_mode,
    coalesce((ns.process_filter ->> 'all_states')::boolean, false) as all_states,
    ns.process_filter -> 'process_states' as process_states,
    nullif(ns.process_filter ->> 'include_user_id', '') as include_user_id
  from lca_active_snapshots act
  join lca_network_snapshots ns on ns.id = act.snapshot_id
  order by act.activated_at desc
  limit 1
),
scope_processes as (
  select
    p.id as process_id,
    p.version as process_version,
    p.state_code,
    p.user_id,
    p.model_id,
    p.json::jsonb as process_json
  from processes p
  cross join active_snapshot snap
  where p.json ? 'processDataSet'
    and (
      snap.all_states
      or p.state_code in (
        select value::int
        from jsonb_array_elements_text(coalesce(snap.process_states, '[]'::jsonb))
      )
      or (snap.include_user_id is not null and p.user_id::text = snap.include_user_id)
    )
),
scope_process_meta as (
  select
    sp.process_id,
    sp.process_version,
    sp.state_code,
    sp.user_id,
    sp.model_id,
    coalesce(
      sp.process_json #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,1,#text}',
      sp.process_json #>> '{processDataSet,processInformation,dataSetInformation,name,baseName,0,#text}'
    ) as process_name,
    coalesce(
      sp.process_json #>> '{processDataSet,processInformation,geography,locationOfOperationSupplyOrProduction,@location}',
      sp.process_json #>> '{processDataSet,processInformation,geography,subLocationOfOperationSupplyOrProduction,@subLocation}',
      sp.process_json #>> '{processDataSet,processInformation,dataSetInformation,locationOfOperationSupplyOrProduction}'
    ) as location,
    sp.process_json
  from scope_processes sp
),
scope_exchanges as (
  select
    pm.process_id,
    pm.process_version,
    pm.process_name,
    pm.location,
    pm.state_code,
    pm.user_id,
    pm.model_id,
    ex.value #>> '{@dataSetInternalID}' as exchange_internal_id,
    ex.value #>> '{exchangeDirection}' as direction,
    ex.value #>> '{referenceToFlowDataSet,@refObjectId}' as flow_id,
    case
      when ex.value ? 'meanAmount' then 'meanAmount'
      when ex.value ? 'resultingAmount' then 'resultingAmount'
      when ex.value ? 'meanValue' then 'meanValue'
      else ''
    end as input_amount_source,
    coalesce(
      ex.value #>> '{meanAmount}',
      ex.value #>> '{resultingAmount}',
      ex.value #>> '{meanValue}',
      ''
    ) as input_amount_text_raw,
    coalesce(
      ex.value #>> '{generalComment,1,#text}',
      ex.value #>> '{generalComment,0,#text}',
      ex.value #>> '{generalComment,#text}',
      ex.value #>> '{common:generalComment,1,#text}',
      ex.value #>> '{common:generalComment,0,#text}',
      ex.value #>> '{common:generalComment,#text}',
      ''
    ) as input_comment
  from scope_process_meta pm
  cross join lateral jsonb_array_elements(
    case
      when jsonb_typeof(pm.process_json #> '{processDataSet,exchanges,exchange}') = 'array' then pm.process_json #> '{processDataSet,exchanges,exchange}'
      when pm.process_json #> '{processDataSet,exchanges,exchange}' is null then '[]'::jsonb
      else jsonb_build_array(pm.process_json #> '{processDataSet,exchanges,exchange}')
    end
  ) ex(value)
),
scope_output_provider_flows as (
  select distinct se.flow_id
  from scope_exchanges se
  where se.direction = 'Output'
    and se.flow_id ~* '^[0-9a-f-]{36}$'
),
parsed_inputs as (
  select
    snap.snapshot_scope,
    snap.snapshot_id,
    snap.snapshot_activated_at,
    snap.snapshot_provider_rule,
    snap.selection_mode,
    snap.include_user_id,
    se.process_id,
    se.process_version,
    se.process_name,
    se.location,
    se.state_code,
    se.user_id,
    case
      when snap.include_user_id is not null and se.user_id::text = snap.include_user_id then 'private'
      else 'public'
    end as process_partition,
    se.exchange_internal_id as input_exchange_internal_id,
    se.input_amount_source,
    trim(se.input_amount_text_raw) as input_amount_text,
    case
      when translate(trim(se.input_amount_text_raw), ',', '') ~ '^[+-]?([0-9]+(\.[0-9]*)?|\.[0-9]+)([eE][+-]?[0-9]+)?$'
      then translate(trim(se.input_amount_text_raw), ',', '')::double precision
      else null
    end as input_amount_value,
    nullif(trim(se.input_comment), '') as input_comment,
    se.flow_id
  from scope_exchanges se
  cross join active_snapshot snap
  where se.direction = 'Input'
    and se.flow_id ~* '^[0-9a-f-]{36}$'
),
unmatched as (
  select *
  from parsed_inputs pi
  where pi.input_amount_value is not null
    and not exists (
      select 1
      from scope_output_provider_flows pf
      where pf.flow_id = pi.flow_id
    )
)
select
  u.snapshot_scope,
  u.snapshot_id,
  u.snapshot_activated_at,
  u.snapshot_provider_rule,
  u.selection_mode,
  coalesce(u.include_user_id, '') as include_user_id,
  u.process_id,
  u.process_version,
  coalesce(u.process_name, '') as process_name,
  coalesce(u.location, '') as location,
  u.state_code,
  coalesce(u.user_id::text, '') as user_id,
  u.process_partition,
  coalesce(u.input_exchange_internal_id, '') as input_exchange_internal_id,
  u.input_amount_source,
  u.input_amount_text,
  u.input_amount_value,
  coalesce(u.input_comment, '') as input_comment,
  u.flow_id
from unmatched u
order by
  u.state_code,
  coalesce(u.process_name, ''),
  u.flow_id,
  u.input_exchange_internal_id
"""


FLOW_META_BATCH_SQL = r"""
with target_flows as (
  select unnest(%(flow_ids)s::text[]) as flow_id
),
latest_flows as (
  select distinct on (f.id)
    f.id::text as flow_id,
    f.version,
    f.created_at,
    f.json::jsonb as flow_json
  from flows f
  join target_flows tf on tf.flow_id = f.id::text
  order by f.id, f.created_at desc
)
select
  lf.flow_id,
  coalesce(
    lf.flow_json #>> '{flowDataSet,flowInformation,dataSetInformation,name,baseName,1,#text}',
    lf.flow_json #>> '{flowDataSet,flowInformation,dataSetInformation,name,baseName,0,#text}'
  ) as flow_name,
  coalesce(
    lf.flow_json #>> '{flowDataSet,flowProperties,flowProperty,referenceToFlowPropertyDataSet,common:shortDescription,#text}',
    lf.flow_json #>> '{flowDataSet,flowProperties,flowProperty,referenceToFlowPropertyDataSet,shortDescription,#text}'
  ) as flow_reference_property_name,
  true as flow_exists_in_flow_table
from latest_flows lf
"""


SHEET_GUIDE = [
    ("summary", "总览最新 active snapshot 的 unmatched_no_provider 规模、分类和字段说明。"),
    ("unmatched_no_provider", "逐条 Input exchange 明细，可直接用于排查和后续补数。"),
    ("missing_flow_summary", "按缺失 flow 聚合，适合排优先级和识别缺的是 flow 还是 provider。"),
    ("process_summary", "按 process 聚合，适合定位哪些工艺最受影响。"),
]


CATEGORY_GUIDE = [
    (
        "flow_missing_in_flow_table",
        "输入引用的 flow_id 在 flows 表里都不存在。优先检查 flow 主数据缺失、引用错误或导数脏数据。",
    ),
    (
        "no_provider_in_scope",
        "flow 主数据存在，但在当前 active snapshot 的数据网络范围内没有任何 provider。先补 provider process / output exchange，或后续再单独分析是不是 scope 筛选问题。",
    ),
]


FIELD_GUIDE = [
    ("snapshot_id", "本次导出对应的最新 active snapshot ID。"),
    ("snapshot_scope / snapshot_provider_rule", "当前 snapshot 的 scope 和 provider rule，用于确认导出口径。"),
    ("process_id / process_version", "命中的 consumer process 主键和版本，是定位问题数据最重要的字段。"),
    ("process_name / location / state_code", "consumer process 的名称、地理位置和状态码。"),
    ("user_id / process_partition", "process 的 user 归属，以及在 snapshot 里被当作 public 还是 private。"),
    ("input_exchange_internal_id", "process JSON 内输入 exchange 的 internal id，后续精确修数常用。"),
    ("input_amount_source / input_amount_text / input_amount_value", "输入 amount 的来源字段、原始文本和解析后的数值。"),
    ("input_comment", "输入 exchange 的注释文本，便于人工理解语义。"),
    ("flow_id / flow_name / flow_reference_property_name", "被引用但没有 provider 的 flow 标识、名称和参考属性。"),
    ("flow_exists_in_flow_table", "这个 flow_id 在 flows 表里是否存在。为 false 时通常优先查 flow 主数据。"),
    ("missing_provider_category", "这条 unmatched_no_provider 属于哪一类根因。建议先按这个字段筛选。"),
    ("triage_hint", "给排查或补数的第一步建议。"),
]


SUMMARY_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CATEGORY_FILLS = {
    "flow_missing_in_flow_table": PatternFill("solid", fgColor="FCE4D6"),
    "no_provider_in_scope": PatternFill("solid", fgColor="E2F0D9"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export unmatched_no_provider diagnostics from the latest active snapshot to a styled Excel workbook."
    )
    parser.add_argument(
        "--report-dir",
        default="reports/unmatched-no-provider-diagnostics",
        help="Directory to write the Excel workbook into.",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Exact output xlsx path. If omitted, use report-dir + timestamped filename.",
    )
    parser.add_argument(
        "--filename-prefix",
        default="unmatched_no_provider_diagnostics",
        help="Filename prefix when --output is not provided.",
    )
    parser.add_argument(
        "--conn-env",
        default="CONN",
        help="Primary environment variable name for database connection. Falls back to DATABASE_URL.",
    )
    return parser.parse_args()


def load_dotenv(dotenv_path: Path) -> None:
    if not dotenv_path.exists():
        return
    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        os.environ.setdefault(key, value)


def repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def database_url(conn_env_name: str) -> str:
    root = repo_root()
    load_dotenv(root / ".env")
    db_url = os.environ.get(conn_env_name) or os.environ.get("DATABASE_URL", "")
    if not db_url:
        raise SystemExit(
            f"missing database connection; set {conn_env_name} or DATABASE_URL, or provide them in {root / '.env'}"
        )
    return db_url


def query_dataframe(conn: psycopg.Connection, sql: str) -> pd.DataFrame:
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()
        columns = [desc.name for desc in cur.description]
    return pd.DataFrame(rows, columns=columns)


def chunked(items: list[str], size: int) -> list[list[str]]:
    return [items[idx : idx + size] for idx in range(0, len(items), size)]


def fetch_flow_meta_df(
    conn: psycopg.Connection,
    flow_ids: list[str],
    batch_size: int = 400,
) -> pd.DataFrame:
    if not flow_ids:
        return pd.DataFrame(
            columns=[
                "flow_id",
                "flow_name",
                "flow_reference_property_name",
                "flow_exists_in_flow_table",
            ]
        )

    frames: list[pd.DataFrame] = []
    with conn.cursor() as cur:
        for batch in chunked(flow_ids, batch_size):
            cur.execute(FLOW_META_BATCH_SQL, {"flow_ids": batch})
            rows = cur.fetchall()
            columns = [desc.name for desc in cur.description]
            frames.append(pd.DataFrame(rows, columns=columns))
    if not frames:
        return pd.DataFrame()
    flow_df = pd.concat(frames, ignore_index=True)
    if flow_df.empty:
        return flow_df
    return flow_df.drop_duplicates(subset=["flow_id"], keep="first").reset_index(drop=True)


def enrich_detail_df(
    detail_df: pd.DataFrame,
    flow_meta_df: pd.DataFrame,
) -> pd.DataFrame:
    if detail_df.empty:
        for column, default in (
            ("flow_name", ""),
            ("flow_reference_property_name", ""),
            ("flow_exists_in_flow_table", False),
            ("missing_provider_category", ""),
            ("triage_hint", ""),
        ):
            detail_df[column] = default
        return detail_df

    merged = detail_df.merge(flow_meta_df, on="flow_id", how="left")
    merged["flow_name"] = merged["flow_name"].fillna("")
    merged["flow_reference_property_name"] = merged["flow_reference_property_name"].fillna("")
    merged["flow_exists_in_flow_table"] = merged["flow_exists_in_flow_table"].map(
        lambda value: bool(value) if pd.notna(value) else False
    )

    merged["missing_provider_category"] = merged.apply(
        lambda row: (
            "flow_missing_in_flow_table"
            if not bool(row["flow_exists_in_flow_table"])
            else "no_provider_in_scope"
        ),
        axis=1,
    )
    merged["triage_hint"] = merged["missing_provider_category"].map(triage_hint_for_category)
    return merged.sort_values(
        by=[
            "missing_provider_category",
            "state_code",
            "process_name",
            "flow_name",
            "input_exchange_internal_id",
        ],
        ascending=[True, True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)


def output_path_from_args(args: argparse.Namespace) -> Path:
    root = repo_root()
    if args.output:
        return Path(args.output).expanduser().resolve()
    report_dir = (root / args.report_dir).resolve()
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return report_dir / f"{args.filename_prefix}_{ts}.xlsx"


def triage_hint_for_category(category: str) -> str:
    if category == "flow_missing_in_flow_table":
        return "先补 flow 主数据或核对 flow_id 引用是否错误。"
    return "先补能输出该 flow 的 provider process 或 output exchange 数据。"


def build_flow_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    if detail_df.empty:
        return pd.DataFrame(
            columns=[
                "missing_provider_category",
                "triage_hint",
                "flow_id",
                "flow_name",
                "flow_reference_property_name",
                "flow_exists_in_flow_table",
                "unmatched_exchange_count",
                "impacted_process_version_count",
                "impacted_process_id_count",
                "impacted_state_codes",
                "impacted_locations",
            ]
        )

    def join_first(series: pd.Series, limit: int = 5, default: str = "") -> str:
        values = [str(v).strip() for v in series if pd.notna(v) and str(v).strip()]
        if not values:
            return default
        unique = list(dict.fromkeys(values))
        return " | ".join(unique[:limit])

    grouped = (
        detail_df.groupby(
            [
                "missing_provider_category",
                "flow_id",
                "flow_name",
                "flow_reference_property_name",
                "flow_exists_in_flow_table",
            ],
            dropna=False,
        )
        .agg(
            unmatched_exchange_count=("input_exchange_internal_id", "count"),
            impacted_process_version_count=("process_id", lambda s: s.astype(str).str.cat(detail_df.loc[s.index, "process_version"].astype(str), sep="@").nunique()),
            impacted_process_id_count=("process_id", "nunique"),
            impacted_state_codes=("state_code", lambda s: ",".join(map(str, sorted(set(s.tolist()))))),
            impacted_locations=("location", lambda s: join_first(s, limit=5, default="")),
        )
        .reset_index()
    )
    grouped.insert(
        1,
        "triage_hint",
        grouped["missing_provider_category"].map(triage_hint_for_category),
    )
    return grouped.sort_values(
        by=[
            "missing_provider_category",
            "unmatched_exchange_count",
            "impacted_process_version_count",
            "flow_name",
        ],
        ascending=[True, False, False, True],
        kind="stable",
    ).reset_index(drop=True)


def build_process_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    if detail_df.empty:
        return pd.DataFrame(
            columns=[
                "process_id",
                "process_version",
                "process_name",
                "location",
                "state_code",
                "user_id",
                "process_partition",
                "unmatched_exchange_count",
                "distinct_missing_flow_count",
                "flow_missing_in_flow_table_count",
                "no_provider_in_scope_count",
                "top_missing_flows",
                "triage_hint",
            ]
        )

    records: list[dict[str, object]] = []
    for keys, group in detail_df.groupby(
        [
            "process_id",
            "process_version",
            "process_name",
            "location",
            "state_code",
            "user_id",
            "process_partition",
        ],
        dropna=False,
    ):
        counts = group["missing_provider_category"].value_counts()
        top_flows = (
            group.assign(flow_label=group["flow_name"].where(group["flow_name"].astype(str).str.strip() != "", group["flow_id"]))
            .groupby("flow_label")
            .size()
            .sort_values(ascending=False)
        )
        if counts.get("flow_missing_in_flow_table", 0) > 0:
            hint = triage_hint_for_category("flow_missing_in_flow_table")
        else:
            hint = triage_hint_for_category("no_provider_in_scope")
        records.append(
            {
                "process_id": keys[0],
                "process_version": keys[1],
                "process_name": keys[2],
                "location": keys[3],
                "state_code": keys[4],
                "user_id": keys[5],
                "process_partition": keys[6],
                "unmatched_exchange_count": int(len(group)),
                "distinct_missing_flow_count": int(group["flow_id"].nunique()),
                "flow_missing_in_flow_table_count": int(counts.get("flow_missing_in_flow_table", 0)),
                "no_provider_in_scope_count": int(counts.get("no_provider_in_scope", 0)),
                "top_missing_flows": " | ".join(map(str, top_flows.head(5).index.tolist())),
                "triage_hint": hint,
            }
        )
    return pd.DataFrame(records).sort_values(
        by=[
            "unmatched_exchange_count",
            "distinct_missing_flow_count",
            "process_name",
        ],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)


def build_summary_df(snapshot_df: pd.DataFrame, detail_df: pd.DataFrame, flow_summary_df: pd.DataFrame) -> pd.DataFrame:
    snap = snapshot_df.iloc[0].to_dict()
    rows: list[dict[str, object]] = [
        {"项目": "生成时间", "说明": datetime.now().isoformat(timespec="seconds")},
        {"项目": "snapshot_id", "说明": str(snap["snapshot_id"])},
        {"项目": "snapshot_scope", "说明": snap["snapshot_scope"]},
        {"项目": "snapshot_activated_at", "说明": str(snap["snapshot_activated_at"])},
        {"项目": "snapshot_provider_rule", "说明": snap["snapshot_provider_rule"]},
        {"项目": "selection_mode", "说明": snap["selection_mode"]},
        {"项目": "include_user_id", "说明": snap["include_user_id"]},
        {"项目": "process_states_label", "说明": snap["process_states_label"]},
        {"项目": "input_edges_total", "说明": int(snap["input_edges_total"])},
        {"项目": "matched_unique_provider", "说明": int(snap["matched_unique_provider"])},
        {"项目": "matched_multi_provider", "说明": int(snap["matched_multi_provider"])},
        {"项目": "matched_multi_resolved", "说明": int(snap["matched_multi_resolved"])},
        {"项目": "matched_multi_unresolved", "说明": int(snap["matched_multi_unresolved"])},
        {"项目": "matched_multi_fallback_equal", "说明": int(snap["matched_multi_fallback_equal"])},
        {"项目": "unmatched_no_provider_from_snapshot_coverage", "说明": int(snap["unmatched_no_provider"])},
        {"项目": "unmatched_no_provider_detail_rows_exported", "说明": int(len(detail_df))},
        {"项目": "impacted_process_versions", "说明": int(detail_df[["process_id", "process_version"]].drop_duplicates().shape[0]) if not detail_df.empty else 0},
        {"项目": "impacted_process_ids", "说明": int(detail_df["process_id"].nunique()) if not detail_df.empty else 0},
        {"项目": "distinct_missing_flows", "说明": int(detail_df["flow_id"].nunique()) if not detail_df.empty else 0},
    ]

    for category, count in detail_df["missing_provider_category"].value_counts().items():
        rows.append({"项目": f"row_count::{category}", "说明": int(count)})

    if not flow_summary_df.empty:
        for category, count in Counter(flow_summary_df["missing_provider_category"]).items():
            rows.append({"项目": f"flow_count::{category}", "说明": int(count)})

    if not detail_df.empty:
        for state_code, count in Counter(detail_df["state_code"]).most_common():
            rows.append({"项目": f"state_code::{state_code}", "说明": int(count)})

        top_flows = (
            detail_df.assign(flow_label=detail_df["flow_name"].where(detail_df["flow_name"].astype(str).str.strip() != "", detail_df["flow_id"]))
            .groupby("flow_label")
            .size()
            .sort_values(ascending=False)
            .head(10)
        )
        for flow_label, count in top_flows.items():
            rows.append({"项目": f"top_flow::{flow_label}", "说明": int(count)})

    return pd.DataFrame(rows)


def prepare_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    prepared = df.copy()
    for column in prepared.columns:
        series = prepared[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            prepared[column] = series.astype(str)
    return prepared


def autosize_columns(ws, max_width: int = 42) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 10), max_width)


def add_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_summary_sheet(
    path: Path,
    summary_df: pd.DataFrame,
    sheet_guide_df: pd.DataFrame,
    category_guide_df: pd.DataFrame,
    field_guide_df: pd.DataFrame,
) -> None:
    wb = load_workbook(path)
    ws = wb["summary"]
    ws.freeze_panes = "A3"

    ws["A1"] = "unmatched_no_provider 排查总览"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "本表基于最新 active snapshot 重建 unmatched_no_provider 明细，并区分是 flow 主数据缺、provider 数据缺，还是 provider 在 scope 外。"
    ws["A2"].alignment = Alignment(wrap_text=True)

    stats_header_row = 3
    for cell in ws[stats_header_row]:
        cell.fill = SUMMARY_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

    sheet_title_row = len(summary_df) + 6
    sheet_header_row = sheet_title_row + 1
    category_title_row = len(summary_df) + len(sheet_guide_df) + 10
    category_header_row = category_title_row + 1
    field_title_row = len(summary_df) + len(sheet_guide_df) + len(category_guide_df) + 14
    field_header_row = field_title_row + 1

    ws[f"A{sheet_title_row}"] = "Sheet 说明"
    ws[f"A{category_title_row}"] = "分类含义"
    ws[f"A{field_title_row}"] = "关键字段含义"
    for title_cell in (ws[f"A{sheet_title_row}"], ws[f"A{category_title_row}"], ws[f"A{field_title_row}"]):
        title_cell.fill = SECTION_FILL
        title_cell.font = BOLD_FONT

    for header_row in (sheet_header_row, category_header_row, field_header_row):
        for cell in ws[header_row]:
            cell.fill = SUMMARY_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row >= stats_header_row and cell.column <= 2:
                cell.border = THIN_BORDER

    autosize_columns(ws, max_width=68)
    wb.save(path)


def style_data_sheet(path: Path, sheet_name: str, category_column: str | None = None) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    category_idx = header_map.get(category_column) if category_column else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = None
        if category_idx:
            fill = CATEGORY_FILLS.get(str(row[category_idx - 1].value))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    autosize_columns(ws, max_width=42 if sheet_name == "summary" else 36)
    add_table(ws, 1, ws.max_row, ws.max_column, f"{sheet_name.replace('-', '_')}_tbl")
    wb.save(path)


def export_workbook(
    output_path: Path,
    snapshot_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    flow_summary_df: pd.DataFrame,
    process_summary_df: pd.DataFrame,
) -> None:
    summary_df = build_summary_df(snapshot_df, detail_df, flow_summary_df)
    sheet_guide_df = pd.DataFrame(SHEET_GUIDE, columns=["sheet", "中文说明"])
    category_guide_df = pd.DataFrame(CATEGORY_GUIDE, columns=["missing_provider_category", "中文解释"])
    field_guide_df = pd.DataFrame(FIELD_GUIDE, columns=["字段", "中文解释"])
    detail_export_df = prepare_df_for_excel(detail_df)
    flow_summary_export_df = prepare_df_for_excel(flow_summary_df)
    process_summary_export_df = prepare_df_for_excel(process_summary_df)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="summary", startrow=2)
        sheet_guide_df.to_excel(writer, index=False, sheet_name="summary", startrow=len(summary_df) + 6)
        category_guide_df.to_excel(
            writer,
            index=False,
            sheet_name="summary",
            startrow=len(summary_df) + len(sheet_guide_df) + 10,
        )
        field_guide_df.to_excel(
            writer,
            index=False,
            sheet_name="summary",
            startrow=len(summary_df) + len(sheet_guide_df) + len(category_guide_df) + 14,
        )
        detail_export_df.to_excel(writer, index=False, sheet_name="unmatched_no_provider")
        flow_summary_export_df.to_excel(writer, index=False, sheet_name="missing_flow_summary")
        process_summary_export_df.to_excel(writer, index=False, sheet_name="process_summary")

    style_summary_sheet(output_path, summary_df, sheet_guide_df, category_guide_df, field_guide_df)
    style_data_sheet(output_path, "unmatched_no_provider", category_column="missing_provider_category")
    style_data_sheet(output_path, "missing_flow_summary", category_column="missing_provider_category")
    style_data_sheet(output_path, "process_summary")


def main() -> None:
    args = parse_args()
    db_url = database_url(args.conn_env)
    out_path = output_path_from_args(args)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    conn = psycopg.connect(db_url)
    try:
        snapshot_df = query_dataframe(conn, ACTIVE_SNAPSHOT_SQL)
        detail_df = query_dataframe(conn, DETAIL_SQL)
        flow_ids = sorted(detail_df["flow_id"].dropna().astype(str).unique().tolist())
        flow_meta_df = fetch_flow_meta_df(conn, flow_ids)
    finally:
        conn.close()

    if snapshot_df.empty:
        raise SystemExit("no active snapshot found")

    detail_df = enrich_detail_df(detail_df, flow_meta_df)
    flow_summary_df = build_flow_summary(detail_df)
    process_summary_df = build_process_summary(detail_df)
    export_workbook(out_path, snapshot_df, detail_df, flow_summary_df, process_summary_df)

    exported_rows = len(detail_df)
    snapshot_id = snapshot_df.iloc[0]["snapshot_id"]
    expected_rows = int(snapshot_df.iloc[0]["unmatched_no_provider"])
    status = "OK" if exported_rows == expected_rows else "MISMATCH"
    print(f"[done] exported {out_path}")
    print(f"[snapshot_id] {snapshot_id}")
    print(f"[detail_rows] {exported_rows}")
    print(f"[coverage_unmatched_no_provider] {expected_rows}")
    print(f"[check] {status}")


if __name__ == "__main__":
    main()
